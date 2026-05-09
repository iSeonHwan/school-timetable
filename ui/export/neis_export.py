"""
NEIS 내보내기 — Excel(.xlsx) 출력

NEIS(국가교육정보시스템)에 직접 업로드할 수 있는 형식이 아닌,
NEIS 시간표 템플릿에 복사·붙여넣기할 수 있도록 정리된 Excel 파일을 생성합니다.

출력 범위:
  - 반별 시간표: 각 반마다 별도 시트 생성
  - 교사별 시간표: 각 교사마다 별도 시트 생성
  - 모두: 위 두 가지 모두 생성

각 시트 구조:
  1행: 타이틀 (병합 셀)
  2행: 비어 있음 (여백)
  3행: 헤더 (교시 레이블 / 요일)
  4행~: 교시별 수업 데이터 (교과약어 + 교사명)
"""
from PyQt6.QtWidgets import (
    QDialog, QFormLayout, QHBoxLayout,
    QLabel, QComboBox, QPushButton, QFileDialog, QMessageBox,
    QDialogButtonBox,
)
from database.connection import get_session
from database.models import (
    TimetableEntry, SchoolClass, Grade, Teacher, AcademicTerm,
)

DAYS_KR = ["월", "화", "수", "목", "금"]


class NEISExportDialog(QDialog):
    """NEIS Excel 내보내기 설정 다이얼로그."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("NEIS 내보내기")
        self.setMinimumWidth(400)
        self._init_ui()

    def _init_ui(self):
        layout = QFormLayout(self)

        session = get_session()
        try:
            # 학기 콤보박스
            self._cmb_term = QComboBox()
            terms = session.query(AcademicTerm).order_by(
                AcademicTerm.year.desc(), AcademicTerm.semester.desc()
            ).all()
            for t in terms:
                self._cmb_term.addItem(str(t), t.id)
            if not terms:
                self._cmb_term.addItem("(학기 없음)", None)
            layout.addRow("학기:", self._cmb_term)

            # 내보내기 범위 콤보박스
            self._cmb_scope = QComboBox()
            self._cmb_scope.addItem("반별 시간표", "classes")
            self._cmb_scope.addItem("교사별 시간표", "teachers")
            self._cmb_scope.addItem("모두", "both")
            layout.addRow("내보내기 범위:", self._cmb_scope)
        finally:
            session.close()

        # 저장 경로 선택
        path_layout = QHBoxLayout()
        self._lbl_path = QLabel("(선택 안 됨)")
        self._lbl_path.setStyleSheet("color:#888;")
        path_layout.addWidget(self._lbl_path)
        btn_browse = QPushButton("찾아보기...")
        btn_browse.clicked.connect(self._browse)
        path_layout.addWidget(btn_browse)
        layout.addRow("저장 경로:", path_layout)

        # 안내 문구
        note = QLabel(
            "NEIS에 직접 업로드 가능한 형식이 아닌, "
            "NEIS 템플릿에 복사하여 사용할 수 있는 정리된 Excel 파일을 생성합니다."
        )
        note.setWordWrap(True)
        note.setStyleSheet("color:#888; font-size:10pt; padding:8px 0;")
        layout.addRow(note)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._export)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def _browse(self):
        """파일 저장 경로를 선택하는 다이얼로그를 엽니다."""
        path, _ = QFileDialog.getSaveFileName(
            self, "NEIS Excel 저장 경로", "timetable_neis.xlsx",
            "Excel Files (*.xlsx)"
        )
        if path:
            self._lbl_path.setText(path)

    def _export(self):
        """입력 값을 검증하고 export_neis() 를 호출합니다."""
        filepath = self._lbl_path.text()
        if filepath == "(선택 안 됨)":
            QMessageBox.warning(self, "경로 오류", "저장 경로를 선택해 주세요.")
            return

        term_id = self._cmb_term.currentData()
        if not term_id:
            QMessageBox.warning(self, "학기 오류", "학기를 선택해 주세요.")
            return

        scope = self._cmb_scope.currentData()
        session = get_session()
        try:
            export_neis(session, term_id, scope, filepath)
        finally:
            session.close()

        self.accept()


def export_neis(session, term_id: int, scope: str, filepath: str) -> None:
    """
    openpyxl 로 Excel 파일을 생성합니다.

    Args:
        session  : 열린 SQLAlchemy 세션
        term_id  : 출력할 학기 ID
        scope    : "classes" / "teachers" / "both"
        filepath : 저장할 .xlsx 파일 경로
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 공통 스타일 정의 ─────────────────────────────────────────────
    header_fill  = PatternFill(start_color="1B4F8A", end_color="1B4F8A", fill_type="solid")
    header_font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    cell_font    = Font(name="맑은 고딕", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def build_class_sheet(ws, title: str, entries: list) -> None:
        """
        단일 시트에 시간표 그리드를 씁니다.
        행: 교시(1~max_periods), 열: 요일(월~금)
        """
        max_periods = max((e.period for e in entries), default=7)

        # 1행: 타이틀 (A1~F1 병합)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font      = Font(name="맑은 고딕", bold=True, size=14, color="1B4F8A")
        title_cell.alignment = Alignment(horizontal="center")

        # 3행: 헤더 (빈칸 + 요일)
        headers = [""] + DAYS_KR
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border

        # 4행~: 교시별 데이터
        for period in range(1, max_periods + 1):
            row = period + 3

            # 교시 레이블 (A열)
            cell = ws.cell(row=row, column=1, value=f"{period}교시")
            cell.font      = Font(name="맑은 고딕", bold=True, size=9)
            cell.alignment = center_align
            cell.fill      = PatternFill(start_color="E8ECF0", end_color="E8ECF0", fill_type="solid")
            cell.border    = thin_border

            # 요일별 수업 데이터 (B~F열)
            for day in range(1, 6):
                col = day + 1
                # entries 에서 (day, period) 에 해당하는 항목을 찾습니다.
                entry = next(
                    (e for e in entries if e.day_of_week == day and e.period == period),
                    None,
                )
                if entry:
                    subj_name = entry.subject.short_name if entry.subject else ""
                    tchr_name = entry.teacher.name if entry.teacher else ""
                    value = f"{subj_name}\n{tchr_name}"
                else:
                    value = ""
                cell        = ws.cell(row=row, column=col, value=value)
                cell.font   = cell_font
                cell.alignment = center_align
                cell.border = thin_border

        # 열 너비 조정
        ws.column_dimensions["A"].width = 10
        for day in range(5):
            ws.column_dimensions[get_column_letter(day + 2)].width = 18

    # ── 반별 시트 생성 ────────────────────────────────────────────────
    if scope in ("classes", "both"):
        classes = (
            session.query(SchoolClass)
            .join(Grade)
            .order_by(Grade.grade_number, SchoolClass.class_number)
            .all()
        )
        first_class = True
        for cls in classes:
            entries = (
                session.query(TimetableEntry)
                .filter_by(term_id=term_id, school_class_id=cls.id)
                .all()
            )
            if not entries:
                continue
            # 첫 번째 반은 Workbook 기본 시트("Sheet")를 재사용하고,
            # 이후 반은 새 시트를 생성합니다.
            if first_class:
                ws = wb.active
                ws.title = cls.display_name
                first_class = False
            else:
                ws = wb.create_sheet(title=cls.display_name)
            build_class_sheet(ws, f"{cls.display_name} 시간표", entries)

    # ── 교사별 시트 생성 ──────────────────────────────────────────────
    if scope in ("teachers", "both"):
        teachers = session.query(Teacher).order_by(Teacher.name).all()
        first_teacher = True
        for teacher in teachers:
            entries = (
                session.query(TimetableEntry)
                .filter_by(term_id=term_id, teacher_id=teacher.id)
                .all()
            )
            if not entries:
                continue
            safe_name = teacher.name[:20]  # Excel 시트명은 31자 제한
            # scope="teachers" 이고 첫 번째 교사면 기본 시트를 재사용합니다.
            if first_teacher and scope == "teachers":
                ws = wb.active
                ws.title = safe_name
                first_teacher = False
            else:
                ws = wb.create_sheet(title=safe_name)
            build_class_sheet(ws, f"{teacher.name} 선생님 시간표", entries)

    wb.save(filepath)
    QMessageBox.information(
        None, "NEIS 내보내기 완료",
        f"Excel 파일이 저장되었습니다:\n{filepath}\n\n"
        "이 파일을 NEIS 템플릿에 복사하여 사용하세요.",
    )
